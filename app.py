from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import plotly.graph_objs as go
import plotly.utils
import json
import io
import os
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import math
import socket
import tempfile

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('templates', exist_ok=True)
os.makedirs('static', exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

oee_data = []
downtime_data = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def convert_time_to_minutes(time_val):
    """Convertit une durée en minutes"""
    if time_val is None:
        return 0
    try:
        if isinstance(time_val, (dt.time, dt.datetime)):
            return time_val.hour * 60 + time_val.minute + time_val.second / 60
        return 0
    except:
        return 0

def parse_downtime_file(file_path, df_dict):
    """Parse les deux feuilles du fichier downtime via openpyxl"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    downtime_categories = {}
    downtime_causes = {}

    if 'Feuil2' in wb.sheetnames:
        ws = wb['Feuil2']
        for row in ws.iter_rows():
            cells = list(row)
            if len(cells) >= 4:
                label_cell = cells[2]
                dur_cell   = cells[3]
                if label_cell.value and isinstance(label_cell.value, str):
                    label = label_cell.value.strip()
                    keywords = ['setup time', 'no cause time', 'org. time',
                                'fail time', 'break time']
                    if any(k in label.lower() for k in keywords) and dur_cell.value is not None:
                        mins = convert_time_to_minutes(dur_cell.value)
                        if mins > 0:
                            downtime_categories[label] = round(mins, 2)

    if 'Feuil4' in wb.sheetnames:
        ws = wb['Feuil4']
        for row in ws.iter_rows():
            cells = list(row)
            if len(cells) >= 4:
                cause_cell = cells[2]
                dur_cell   = cells[3]
                if cause_cell.value and isinstance(cause_cell.value, str):
                    cause = cause_cell.value.strip()
                    if cause and cause not in ['Downtime', 'duration', 'Downtime ', 'duration ']:
                        mins = convert_time_to_minutes(dur_cell.value)
                        if mins > 0:
                            downtime_causes[cause] = downtime_causes.get(cause, 0) + mins

    causes_list = sorted(
        [{'cause': c, 'duration': round(d, 2)} for c, d in downtime_causes.items()],
        key=lambda x: x['duration'], reverse=True
    )

    return {
        'categories':       downtime_categories,
        'causes':           causes_list,
        'total_categories': round(sum(downtime_categories.values()), 2),
        'total_causes':     round(sum(c['duration'] for c in causes_list), 2)
    }

def detect_and_convert(value):
    """
    Convertit une valeur en pourcentage.
    Toutes les valeurs du fichier sont en format décimal (0.xx ou 1.xx).
    On multiplie systématiquement par 100 pour obtenir le vrai pourcentage.
    """
    if value is None:
        return 0.0
    try:
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        if isinstance(value, str):
            v = value.strip().replace(',', '.')
            if '%' in v:
                return round(float(v.replace('%', '')), 2)
            num = float(v)
        else:
            num = float(value)

        return round(num * 100, 2)
    except:
        return 0.0

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload/oee', methods=['POST'])
def upload_oee():
    global oee_data

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier uploadé'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400

    if file and allowed_file(file.filename):
        try:
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file, encoding='utf-8')
            else:
                df = pd.read_excel(file, engine='openpyxl', sheet_name=0)

            df.columns = [str(col).strip() for col in df.columns]

            print("=== COLONNES TROUVÉES ===")
            for i, col in enumerate(df.columns):
                print(f"  Colonne {i}: '{col}'")

            machine_col      = df.columns[0] if len(df.columns) > 0 else None
            oee_col          = df.columns[1] if len(df.columns) > 1 else None
            availability_col = df.columns[2] if len(df.columns) > 2 else None
            performance_col  = df.columns[3] if len(df.columns) > 3 else None
            quality_col      = df.columns[4] if len(df.columns) > 4 else None
            loading_col      = df.columns[5] if len(df.columns) > 5 else None

            machines_data = []

            for idx, row in df.iterrows():
                machine = str(row[machine_col]).strip() if machine_col and pd.notna(row[machine_col]) else ""

                if machine == '' or machine == 'nan' or machine.lower() in ['machine', 'nom', 'name', 'machine ']:
                    continue

                oee_raw          = row[oee_col]          if oee_col          and pd.notna(row[oee_col])          else 0
                availability_raw = row[availability_col] if availability_col and pd.notna(row[availability_col]) else 0
                performance_raw  = row[performance_col]  if performance_col  and pd.notna(row[performance_col])  else 0
                quality_raw      = row[quality_col]      if quality_col      and pd.notna(row[quality_col])      else 0
                loading_raw      = row[loading_col]      if loading_col      and pd.notna(row[loading_col])      else 0

                oee_pct          = detect_and_convert(oee_raw)
                availability_pct = detect_and_convert(availability_raw)
                performance_pct  = detect_and_convert(performance_raw)
                quality_pct      = detect_and_convert(quality_raw)
                loading_pct      = detect_and_convert(loading_raw)

                machines_data.append({
                    'Machine':      machine,
                    'OEE':          round(oee_pct, 2),
                    'Availability': round(availability_pct, 2),
                    'Performance':  round(performance_pct, 2),
                    'Quality':      round(quality_pct, 2),
                    'Loading':      round(loading_pct, 2)
                })

            seen = set()
            unique_machines = []
            for m in machines_data:
                if m['Machine'] not in seen:
                    seen.add(m['Machine'])
                    unique_machines.append(m)

            oee_data = sorted(unique_machines, key=lambda x: x['OEE'], reverse=True)

            total_machines = len(oee_data)
            if total_machines == 0:
                return jsonify({'error': 'Aucune donnée valide trouvée'}), 400

            print(f"\n=== MACHINES CHARGÉES ({total_machines}) ===")
            for m in oee_data[:5]:
                print(f"  {m['Machine']}: OEE={m['OEE']}%")

            oee_values          = [m['OEE']          for m in oee_data if m['OEE']          > 0]
            availability_values = [m['Availability'] for m in oee_data if m['Availability'] > 0]
            performance_values  = [m['Performance']  for m in oee_data if m['Performance']  > 0]
            quality_values      = [m['Quality']      for m in oee_data if m['Quality']      > 0]
            loading_values      = [m['Loading']      for m in oee_data if m['Loading']      > 0]

            avg_oee          = round(np.mean(oee_values),          2) if oee_values          else 0
            avg_availability = round(np.mean(availability_values), 2) if availability_values else 0
            avg_performance  = round(np.mean(performance_values),  2) if performance_values  else 0
            avg_quality      = round(np.mean(quality_values),      2) if quality_values      else 0
            avg_loading      = round(np.mean(loading_values),      2) if loading_values      else 0

            if oee_values:
                best_idx     = np.argmax([m['OEE'] for m in oee_data])
                best_machine = oee_data[best_idx]['Machine']
                best_oee     = oee_data[best_idx]['OEE']

                non_zero = [(i, m) for i, m in enumerate(oee_data) if m['OEE'] > 0]
                if non_zero:
                    worst_idx_local  = np.argmin([m['OEE'] for _, m in non_zero])
                    worst_machine    = non_zero[worst_idx_local][1]['Machine']
                    worst_oee        = non_zero[worst_idx_local][1]['OEE']
                else:
                    worst_machine = "N/A"
                    worst_oee     = 0

                max_oee         = max(m['OEE'] for m in oee_data)
                max_oee_machine = next(m['Machine'] for m in oee_data if m['OEE'] == max_oee)
            else:
                best_machine    = "N/A"
                best_oee        = 0
                worst_machine   = "N/A"
                worst_oee       = 0
                max_oee         = 0
                max_oee_machine = "N/A"

            oee_red    = len([m for m in oee_data if 0 < m['OEE'] < 55])
            oee_orange = len([m for m in oee_data if 55 <= m['OEE'] < 75])
            oee_green  = len([m for m in oee_data if m['OEE'] >= 75])

            metrics = {
                'total_machines':  total_machines,
                'avg_oee':         avg_oee,
                'avg_availability': avg_availability,
                'avg_performance': avg_performance,
                'avg_quality':     avg_quality,
                'avg_loading':     avg_loading,
                'avg_machine':     avg_oee,
                'best_machine':    best_machine,
                'best_oee':        best_oee,
                'worst_machine':   worst_machine,
                'worst_oee':       worst_oee,
                'max_oee':         max_oee,
                'max_oee_machine': max_oee_machine,
                'oee_red':         oee_red,
                'oee_orange':      oee_orange,
                'oee_green':       oee_green
            }

            graphs = create_oee_graphs(oee_data, metrics)

            return jsonify({
                'success':        True,
                'type':           'oee',
                'metrics':        metrics,
                'graphs':         graphs,
                'machines_data':  oee_data,
                'machines_count': total_machines
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Erreur: {str(e)}'}), 500

    return jsonify({'error': 'Type de fichier non supporté'}), 400

@app.route('/upload/downtime', methods=['POST'])
def upload_downtime():
    global downtime_data

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier uploadé'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400

    if file and allowed_file(file.filename):
        try:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)

            downtime_data = parse_downtime_file(file_path, {})
            graphs = create_downtime_graphs(downtime_data)

            return jsonify({
                'success':       True,
                'type':          'downtime',
                'downtime_data': downtime_data,
                'graphs':        graphs
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Erreur: {str(e)}'}), 500

    return jsonify({'error': 'Type de fichier non supporté'}), 400

def create_oee_graphs(machines_data, metrics):
    if not machines_data:
        empty_fig = go.Figure()
        empty_fig.add_annotation(text="Aucune donnée disponible", xref="paper", yref="paper",
                                  x=0.5, y=0.5, showarrow=False)
        empty_json = json.loads(json.dumps(empty_fig, cls=plotly.utils.PlotlyJSONEncoder))
        return {k: empty_json for k in ['top10', 'bottom10', 'averages', 'gauge', 'performance_chart', 'all_oee']}

    def get_oee_color(oee):
        if oee < 55:   return '#e53e3e'
        elif oee < 75: return '#ed8936'
        else:          return '#48bb78'

    valid_machines = [m for m in machines_data if m['OEE'] > 0] or machines_data
    sorted_by_oee  = sorted(valid_machines, key=lambda x: x['OEE'], reverse=True)

    max_oee_value = max(m['OEE'] for m in valid_machines)
    y_max_oee     = max(120, max_oee_value * 1.1)

    top10 = sorted_by_oee[:10]
    fig_top10 = go.Figure(data=[go.Bar(
        x=[m['Machine'] for m in top10],
        y=[m['OEE']     for m in top10],
        marker_color=[get_oee_color(m['OEE']) for m in top10],
        text=[f"{m['OEE']:.2f}%"             for m in top10],
        textposition='auto'
    )])
    fig_top10.update_layout(
        title="Top 10 Machines - Meilleur OEE",
        xaxis_title="Machine", yaxis_title="OEE (%)",
        height=400, xaxis_tickangle=45, yaxis_range=[0, y_max_oee], template='plotly_white'
    )

    bottom10 = sorted_by_oee[-10:]
    fig_bottom10 = go.Figure(data=[go.Bar(
        x=[m['Machine'] for m in bottom10],
        y=[m['OEE']     for m in bottom10],
        marker_color=[get_oee_color(m['OEE']) for m in bottom10],
        text=[f"{m['OEE']:.2f}%"             for m in bottom10],
        textposition='auto'
    )])
    fig_bottom10.update_layout(
        title="Bottom 10 Machines - À améliorer",
        xaxis_title="Machine", yaxis_title="OEE (%)",
        height=400, xaxis_tickangle=45, template='plotly_white'
    )

    fig_avg = go.Figure(data=[
        go.Bar(name='OEE',          x=['Moyennes'], y=[metrics['avg_oee']],          marker_color=get_oee_color(metrics['avg_oee'])),
        go.Bar(name='Availability', x=['Moyennes'], y=[metrics['avg_availability']], marker_color='#4ECDC4'),
        go.Bar(name='Performance',  x=['Moyennes'], y=[metrics['avg_performance']],  marker_color='#45B7D1'),
        go.Bar(name='Quality',      x=['Moyennes'], y=[metrics['avg_quality']],      marker_color='#96CEB4'),
        go.Bar(name='Loading',      x=['Moyennes'], y=[metrics['avg_loading']],      marker_color='#FFEAA7'),
    ])
    fig_avg.update_layout(
        title="Moyennes Globales", yaxis_title="Pourcentage (%)",
        height=400, template='plotly_white'
    )

    display_oee = min(metrics['avg_oee'], 100)
    fig_gauge = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=display_oee,
        number={'suffix': '%', 'valueformat': '.2f'},
        title={'text': f"OEE Moyen Global<br><sub>(moyenne réelle: {metrics['avg_oee']:.2f}%)</sub>",
               'font': {'size': 14}},
        delta={'reference': 75},
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': get_oee_color(metrics['avg_oee'])},
            'steps': [
                {'range': [0,  55], 'color': '#e53e3e'},
                {'range': [55, 75], 'color': '#ed8936'},
                {'range': [75,100], 'color': '#48bb78'},
            ],
            'threshold': {'line': {'color': 'red', 'width': 4}, 'value': 75}
        }
    ))
    fig_gauge.update_layout(height=350)

    sorted_by_perf = sorted(valid_machines, key=lambda x: x['Performance'], reverse=True)
    top10_perf     = sorted_by_perf[:10]
    max_perf_value = max(m['Performance'] for m in valid_machines)
    y_max_perf     = max(120, max_perf_value * 1.1)

    fig_perf = go.Figure(data=[go.Bar(
        x=[m['Machine']     for m in top10_perf],
        y=[m['Performance'] for m in top10_perf],
        marker_color='#FFB347',
        text=[f"{m['Performance']:.2f}%" for m in top10_perf],
        textposition='auto'
    )])
    fig_perf.update_layout(
        title="Top 10 Performance Machine",
        xaxis_title="Machine", yaxis_title="Performance (%)",
        height=400, xaxis_tickangle=45, yaxis_range=[0, y_max_perf], template='plotly_white'
    )

    fig_all_oee = go.Figure(data=[go.Bar(
        x=[m['Machine'] for m in sorted_by_oee],
        y=[m['OEE']     for m in sorted_by_oee],
        marker_color=[get_oee_color(m['OEE']) for m in sorted_by_oee],
        text=[f"{m['OEE']:.2f}%"             for m in sorted_by_oee],
        textposition='auto', textangle=0
    )])
    fig_all_oee.update_layout(
        title="Toutes les machines - Classement OEE",
        xaxis_title="Machine", yaxis_title="OEE (%)",
        height=500, xaxis_tickangle=90, yaxis_range=[0, y_max_oee], template='plotly_white'
    )

    def to_json(fig):
        return json.loads(json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder))

    return {
        'top10':             to_json(fig_top10),
        'bottom10':          to_json(fig_bottom10),
        'averages':          to_json(fig_avg),
        'gauge':             to_json(fig_gauge),
        'performance_chart': to_json(fig_perf),
        'all_oee':           to_json(fig_all_oee),
    }

def create_downtime_graphs(downtime_data):
    """Crée les graphiques pour les temps d'arrêt"""
    graphs = {}

    if downtime_data.get('categories'):
        categories = downtime_data['categories']
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']

        fig_pie = go.Figure(data=[go.Pie(
            labels=list(categories.keys()),
            values=list(categories.values()),
            hole=0.4,
            marker=dict(colors=colors[:len(categories)]),
            textinfo='label+percent+value'
        )])
        fig_pie.update_layout(
            title="Distribution des catégories d'arrêt",
            height=450, template='plotly_white'
        )
        graphs['categories_pie'] = json.loads(json.dumps(fig_pie, cls=plotly.utils.PlotlyJSONEncoder))

    if downtime_data.get('causes'):
        causes = downtime_data['causes'][:30]
        causes_names = []
        for c in causes:
            name = c['cause']
            if len(name) > 40:
                name = name[:37] + '...'
            causes_names.append(name)

        causes_durations = [c['duration'] for c in causes]

        fig_causes = go.Figure(data=[go.Bar(
            x=causes_names,
            y=causes_durations,
            marker_color='#FF8C42',
            text=[f"{v:.0f} min" for v in causes_durations],
            textposition='outside',
            orientation='v'
        )])
        fig_causes.update_layout(
            title=f"Toutes les causes d'arrêt ({len(causes)} causes)",
            xaxis_title="Cause",
            yaxis_title="Durée (minutes)",
            height=650,
            xaxis_tickangle=45,
            margin=dict(l=50, r=50, t=80, b=180),
            template='plotly_white'
        )
        graphs['causes_bar'] = json.loads(json.dumps(fig_causes, cls=plotly.utils.PlotlyJSONEncoder))

    return graphs

@app.route('/export/oee/csv', methods=['POST'])
def export_oee_csv():
    """Exporte les données OEE au format CSV"""
    data = request.json.get('data', [])
    if not data:
        return jsonify({'error': 'Aucune donnée à exporter'}), 400
    
    df = pd.DataFrame(data)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    temp_file.close()
    
    df.to_csv(temp_file.name, index=False, encoding='utf-8-sig')
    
    return send_file(
        temp_file.name,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'oee_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@app.route('/export/oee/excel', methods=['POST'])
def export_oee_excel():
    """Exporte les données OEE au format Excel"""
    data = request.json.get('data', [])
    if not data:
        return jsonify({'error': 'Aucune donnée à exporter'}), 400
    
    df = pd.DataFrame(data)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='OEE_Data', index=False)
    
    return send_file(
        temp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'oee_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/export/downtime/csv', methods=['POST'])
def export_downtime_csv():
    """Exporte les données Downtime au format CSV"""
    data = request.json.get('data', [])
    if not data:
        return jsonify({'error': 'Aucune donnée à exporter'}), 400
    
    df = pd.DataFrame(data)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    temp_file.close()
    
    df.to_csv(temp_file.name, index=False, encoding='utf-8-sig')
    
    return send_file(
        temp_file.name,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'downtime_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@app.route('/export/downtime/excel', methods=['POST'])
def export_downtime_excel():
    """Exporte les données Downtime au format Excel"""
    data = request.json.get('data', [])
    if not data:
        return jsonify({'error': 'Aucune donnée à exporter'}), 400
    
    df = pd.DataFrame(data)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Downtime_Data', index=False)
    
    return send_file(
        temp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'downtime_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

if __name__ == '__main__':
    import socket
    
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    print("=" * 70)
    print("🚀 Dashboard OEE & Downtime - C&C Area")
    print("=" * 70)
    print(f"📁 Accès local (sur ce PC) : http://localhost:5000")
    print(f"📁 Accès réseau (partage)  : http://{local_ip}:5000")
    print("=" * 70)
    print("⚠️  INSTRUCTIONS POUR VOS COLLÈGUES :")
    print(f"   1. Donnez-leur le lien : http://{local_ip}:5000")
    print("   2. Ils doivent être sur le MÊME RÉSEAU (Wi-Fi ou Ethernet)")
    print("   3. Si ça ne fonctionne pas, désactivez le pare-feu Windows")
    print("=" * 70)
    print("💡 Pour arrêter le serveur : CTRL + C")
    print("=" * 70)
    
    app.run(debug=False, host='0.0.0.0', port=5000)